from __future__ import annotations

import io
import os
import re
import time
import unicodedata
from typing import Any, Optional

import openpyxl

DEFAULT_TYPE_OPTIONS: dict[str, dict[str, list[str]]] = {
    "BTS": {
        "wallSystem": ["Cast in Place Concrete", "Concrete Tilt Panels", "Masonry", "Wood"],
        "roof": ["Steel", "Wood", "Hybrid"],
        "slab": ["Slab on Grade", "Structural Slab", "Elevated Slab"],
        "foundation": ["Spread Footing", "Piers"],
    },
    "TI": {
        "wallSystem": ["Concrete", "Masonry", "Wood"],
        "roof": ["Steel", "Wood", "Hybrid", "Concrete Slab"],
        "slab": ["Slab on Grade", "Structural Slab", "Elevated Slab"],
        "foundation": [],
    },
}

_SIGNALS = {
    "projecttype", "projectname", "projectnumber", "walls", "wallsystem",
    "roof", "slab", "foundation", "foundations",
}


# ── Text normalisation (mirrors the JS helpers) ───────────────────────────────

def _normalize_text(value: Any) -> str:
    s = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def _normalize_domain(value: Any) -> str:
    s = _normalize_text(value)
    s = re.sub(r"\bcost in place\b", "cast in place", s)
    s = re.sub(r"\bcip\b", "cast in place", s)
    s = re.sub(r"\btitl\b", "tilt", s)
    s = re.sub(r"\bpanels?\b", "panel", s)
    s = re.sub(r"\bpiers?\b", "pier", s)
    s = re.sub(r"\bfootings?\b", "footing", s)
    return s


def _normalize_header(value: Any) -> str:
    return _normalize_text(value).replace(" ", "")


def _canonicalize_type(value: Any) -> str:
    n = _normalize_text(value)
    if not n:
        return ""
    tokens = set(n.split())
    if "bts" in tokens:
        return "BTS"
    if "ti" in tokens:
        return "TI"
    return str(value).strip()


# ── Field matching (mirrors matchesField in JS) ───────────────────────────────

def _matches(row_value: Any, selected: str, field: str) -> bool:
    if not selected:
        return True
    rv = str(row_value or "").strip()
    if not rv:
        return False
    if field == "type":
        a = _normalize_text(_canonicalize_type(rv))
        b = _normalize_text(_canonicalize_type(selected))
    else:
        a = _normalize_domain(rv)
        b = _normalize_domain(selected)
    if not a or not b:
        return False
    return a == b or b in a


# ── Column detection (mirrors detectColumnsFromHeaders in JS) ─────────────────

def detect_columns(headers: list[str]) -> dict[str, Optional[str]]:
    candidates = [(h, _normalize_header(h)) for h in headers]

    def pick(tests: list) -> Optional[str]:
        for test in tests:
            for raw, n in candidates:
                if test(n):
                    return raw
        return None

    return {
        "name": pick([
            lambda n: n == "projectname",
            lambda n: n == "name",
            lambda n: ("project" in n and "name" in n),
            lambda n: ("job" in n and "name" in n),
            lambda n: "projecttitle" in n,
        ]),
        "id": pick([
            lambda n: n == "projectid",
            lambda n: n == "projectnumber",
            lambda n: n == "jobnumber",
            lambda n: n == "id",
            lambda n: ("project" in n and n.endswith("id")),
        ]),
        "type": pick([
            lambda n: n == "projecttype",
            lambda n: ("project" in n and "type" in n),
            lambda n: n == "type",
        ]),
        "wallSystem": pick([
            lambda n: n == "wallsystem",
            lambda n: ("wall" in n and "system" in n),
            lambda n: n == "walls",
            lambda n: n == "wall",
        ]),
        "roof": pick([
            lambda n: n == "roof",
            lambda n: "roof" in n,
        ]),
        "slab": pick([
            lambda n: n == "slab",
            lambda n: "slab" in n,
        ]),
        "foundation": pick([
            lambda n: n == "foundation",
            lambda n: n == "foundations",
            lambda n: "found" in n,
        ]),
        "company": pick([
            lambda n: n == "company",
            lambda n: "company" in n,
            lambda n: n == "firm",
            lambda n: n == "client",
        ]),
    }


# ── Excel sheet selection (mirrors pickBestExcelSheet in JS) ──────────────────

def _pick_best_sheet(wb: openpyxl.Workbook) -> Optional[dict]:
    best: Optional[dict] = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=40, values_only=True))

        best_header_row = 0
        best_header_score = -1
        non_empty_rows = 0

        for r_idx, row in enumerate(rows):
            normalized = [_normalize_header(c) for c in row if c is not None]
            populated = sum(1 for n in normalized if n)
            if populated:
                non_empty_rows += 1
            matched = sum(1 for sig in _SIGNALS if sig in normalized)
            score = matched * 100 + populated
            if score > best_header_score:
                best_header_score = score
                best_header_row = r_idx

        name_lower = sheet_name.lower()
        boost = 5000 if "project log" in name_lower else (1500 if "project" in name_lower else 0)
        overall = best_header_score * 1000 + non_empty_rows * 10 + boost

        if best is None or overall > best["overall"]:
            best = {
                "sheet_name": sheet_name,
                "header_row": best_header_row,
                "header_score": best_header_score,
                "overall": overall,
            }

    return best


# ── Config sheet type option extraction ───────────────────────────────────────

def _extract_config_options(wb: openpyxl.Workbook) -> Optional[dict]:
    config_name = next((n for n in wb.sheetnames if _normalize_text(n) == "config"), None)
    if not config_name:
        return None

    ws = wb[config_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None

    headers = [str(h or "").strip() for h in rows[0]]
    norm_h = [_normalize_header(h) for h in headers]

    def pick_idx(tests: list) -> Optional[int]:
        for test in tests:
            for i, n in enumerate(norm_h):
                if test(n):
                    return i
        return None

    comp_idx = pick_idx([lambda n: n == "componentname", lambda n: n == "component", lambda n: n == "name"])
    app_idx  = pick_idx([lambda n: n == "applicability", lambda n: n == "type", lambda n: "applic" in n])
    cat_idx  = pick_idx([lambda n: n == "category", lambda n: "category" in n])

    if any(x is None for x in (comp_idx, app_idx, cat_idx)):
        return None

    def cat_to_field(cat: str) -> Optional[str]:
        if cat in ("wall", "walls"):
            return "wallSystem"
        if cat == "roof":
            return "roof"
        if cat == "slab":
            return "slab"
        if cat in ("foundation", "foundations"):
            return "foundation"
        return None

    options: dict[str, dict[str, list]] = {
        "BTS": {"wallSystem": [], "roof": [], "slab": [], "foundation": []},
        "TI":  {"wallSystem": [], "roof": [], "slab": [], "foundation": []},
    }

    for row in rows[1:]:
        raw_name = str(row[comp_idx] if comp_idx < len(row) else "").strip()
        if not raw_name:
            continue
        app = _normalize_text(row[app_idx] if app_idx < len(row) else "")
        cat = _normalize_text(row[cat_idx] if cat_idx < len(row) else "")
        field = cat_to_field(cat)
        if not field:
            continue
        cleaned = re.sub(r"\s*\(roof\)\s*", "", raw_name, flags=re.I).strip()
        targets = {"both": ["BTS", "TI"], "bts": ["BTS"], "ti": ["TI"]}.get(app, [])
        for t in targets:
            if cleaned not in options[t][field]:
                options[t][field].append(cleaned)

    has_any = any(
        options[t][f]
        for t in ("BTS", "TI")
        for f in ("wallSystem", "roof", "slab", "foundation")
    )
    return options if has_any else None


# ── Row misalignment fix (mirrors fixCommonRowMisalignment in JS) ─────────────

def _fix_misalignment(rows: list[dict], col_map: dict, type_options: dict) -> list[dict]:
    slab_key = col_map.get("slab")
    found_key = col_map.get("foundation")
    if not slab_key or not found_key:
        return rows

    slab_set = {
        _normalize_domain(v)
        for t in type_options.values()
        for v in t.get("slab", [])
    }
    found_set = {
        _normalize_domain(v)
        for t in type_options.values()
        for v in t.get("foundation", [])
    } | {_normalize_domain("Pier")}

    result = []
    for row in rows:
        sv = str(row.get(slab_key) or "").strip()
        fv = str(row.get(found_key) or "").strip()
        sn = _normalize_domain(sv)
        fn = _normalize_domain(fv)

        if not sv and fv and fn in slab_set:
            row = {**row, slab_key: fv, found_key: ""}
        elif not fv and sv and sn in found_set:
            row = {**row, slab_key: "", found_key: sv}
        elif sv and fv and sn in found_set and fn in slab_set:
            row = {**row, slab_key: fv, found_key: sv}
        result.append(row)

    return result


# ── Excel parsing ─────────────────────────────────────────────────────────────

def parse_excel_bytes(buf: bytes) -> dict:
    # Pass 1 — config sheet
    wb1 = openpyxl.load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
    config_options = _extract_config_options(wb1)
    wb1.close()
    type_options = config_options or DEFAULT_TYPE_OPTIONS

    # Pass 2 — detect best sheet
    wb2 = openpyxl.load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
    best = _pick_best_sheet(wb2)
    wb2.close()

    if not best:
        return {"headers": [], "rows": [], "type_options": type_options, "sheet_name": ""}

    # Pass 3 — read data
    wb3 = openpyxl.load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
    ws = wb3[best["sheet_name"]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb3.close()

    header_idx = best["header_row"]
    raw_headers = [str(c).strip() if c is not None else "" for c in (all_rows[header_idx] or [])]
    headers = [h if h else f"Column {i + 1}" for i, h in enumerate(raw_headers)]

    data_rows = []
    for row in all_rows[header_idx + 1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        obj: dict[str, str] = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            obj[h] = str(val).strip() if val is not None else ""
        data_rows.append(obj)

    return {
        "headers": headers,
        "rows": data_rows,
        "type_options": type_options,
        "sheet_name": best["sheet_name"],
    }


# ── Data source ───────────────────────────────────────────────────────────────

def _download_from_supabase() -> bytes:
    from supabase import create_client
    url = os.environ["SUPABASE_URL"]
    key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ["SUPABASE_KEY"]
    bucket = os.environ["AVS_PROJECT_FINDER_BUCKET"]
    object_path = os.environ["AVS_PROJECT_FINDER_OBJECT_PATH"]
    return create_client(url, key).storage.from_(bucket).download(object_path)


def _read_local() -> bytes:
    path = os.environ["AVS_PROJECT_DB_PATH"]
    with open(path, "rb") as f:
        return f.read()


def _has_supabase_config() -> bool:
    return bool(
        os.environ.get("SUPABASE_URL")
        and (os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_KEY"))
        and os.environ.get("AVS_PROJECT_FINDER_BUCKET")
        and os.environ.get("AVS_PROJECT_FINDER_OBJECT_PATH")
    )


# ── In-memory cache (5-minute TTL) ───────────────────────────────────────────

_cache: dict[str, Any] = {"data": None, "ts": 0.0}
_CACHE_TTL = 300.0


def invalidate_cache() -> None:
    _cache["data"] = None
    _cache["ts"] = 0.0


def get_projects() -> dict:
    now = time.monotonic()
    if _cache["data"] is not None and (now - _cache["ts"]) < _CACHE_TTL:
        return _cache["data"]

    buf = _download_from_supabase() if _has_supabase_config() else _read_local()
    parsed = parse_excel_bytes(buf)
    col_map = detect_columns(parsed["headers"])
    rows = _fix_misalignment(parsed["rows"], col_map, parsed["type_options"])

    data: dict[str, Any] = {
        "headers": parsed["headers"],
        "col_map": col_map,
        "type_options": parsed["type_options"],
        "rows": rows,
        "sheet_name": parsed["sheet_name"],
        "total": len(rows),
    }
    _cache["data"] = data
    _cache["ts"] = now
    return data


# ── Search ────────────────────────────────────────────────────────────────────

def search_projects(filters: dict[str, str], limit: int = 500) -> dict:
    data = get_projects()
    col_map = data["col_map"]
    matches = []

    for row in data["rows"]:
        ok = True
        for field, selected in filters.items():
            if not selected:
                continue
            key = col_map.get(field)
            if not key:
                continue
            if not _matches(row.get(key), selected, field):
                ok = False
                break
        if ok:
            matches.append(row)
            if len(matches) >= limit:
                break

    return {
        "ok": True,
        "headers": data["headers"],
        "col_map": col_map,
        "type_options": data["type_options"],
        "total": data["total"],
        "returned": len(matches),
        "truncated": len(matches) >= limit and data["total"] > limit,
        "rows": matches,
    }
